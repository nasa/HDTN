from openpyxl import Workbook, load_workbook
import os.path

def worksheetName():
    titleOfWorksheet = input("Name of Worksheet --> ")
    return titleOfWorksheet
def header():
    worksheet.append(["Payload Only Rate (mbps)", "Total Rate (mbps)", "Bundles Per Second"])

def fileParser():
    for line in lines: # Reads Each line from file
        line = line.split()
        if line[0] == "Payload" and len(line) == 11:
            worksheet.append([float(line[3]), float(line[7]), float(line[9])])

    
file = open("output.txt", "r") # Opens file
lines = file.readlines() # Reads All lines from file
file.close() # closes file

excelWorkbookName = "bpSinkOutputs.xlsx"

if os.path.isfile(excelWorkbookName) == True:
    workbook = load_workbook(excelWorkbookName)
    worksheet = workbook.create_sheet(worksheetName())
    header()
    fileParser()
else:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheetName()
    header()
    fileParser()
    
workbook.save(excelWorkbookName)
print("Finished exporting data to excel") 
